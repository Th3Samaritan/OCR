"""Extract markdown pipe-tables from OCR output → JSON (for preview) or .xlsx."""
from __future__ import annotations

import io
import re

from openpyxl import Workbook

_SEP = re.compile(r"^:?-{2,}:?$")  # a markdown table separator cell like ---, :--, --:


def _is_separator_row(cells: list[str]) -> bool:
    nonempty = [c.strip() for c in cells if c.strip()]
    return bool(nonempty) and all(_SEP.match(c) for c in nonempty)


def _block_to_rows(block: list[str]) -> list[list[str]] | None:
    if len(block) < 2:
        return None
    rows: list[list[str]] = []
    for line in block:
        s = line.strip()
        if s.startswith("|"):
            s = s[1:]
        if s.endswith("|"):
            s = s[:-1]
        cells = [c.strip() for c in s.split("|")]
        if _is_separator_row(cells):
            continue
        rows.append(cells)
    return rows or None


def extract_tables(markdown: str) -> list[list[list[str]]]:
    """Return every pipe-table as a list of rows (each row a list of cells)."""
    tables, block = [], []
    for line in (markdown or "").splitlines():
        if line.strip().startswith("|"):
            block.append(line)
        else:
            rows = _block_to_rows(block)
            if rows:
                tables.append(rows)
            block = []
    rows = _block_to_rows(block)
    if rows:
        tables.append(rows)
    return tables


def tables_as_json(markdown: str) -> list[dict]:
    out = []
    for i, rows in enumerate(extract_tables(markdown), 1):
        out.append({"name": f"Table {i}", "headers": rows[0], "rows": rows[1:]})
    return out


def tables_to_xlsx(markdown: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    tabs = extract_tables(markdown)
    if not tabs:
        ws = wb.create_sheet("No tables")
        ws["A1"] = "No tables detected in the document."
    for i, rows in enumerate(tabs, 1):
        ws = wb.create_sheet(f"Table{i}"[:31])
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
